# exportar.py — Generación del archivo Excel de cubicación.
# Modificar aquí si cambia el formato de la hoja, las fórmulas Excel, columnas o estilos.
# Las referencias de parámetros en las fórmulas son: B13=largo_carpa, E9=dist_plantas,
# G8=merma_hil(%), G9=merma_trans(%), R2=ancho_carpa, R10=largo_transversal.

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Helpers de estilo ─────────────────────────────────────────────────────────

_THIN = Side(style="thin")
_BRD  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_AZUL = "D9E1F2"
_AMAR = "FFF2CC"


def _c(ws, ref_o_fila, col=None, valor=None,
        negrita=False, centrado=False, borde=False, fondo=None):
    """Escribe en una celda y aplica estilos opcionales."""
    cell = ws[ref_o_fila] if col is None else ws.cell(ref_o_fila, col, valor)
    if col is None and valor is not None:
        cell.value = valor
    if negrita:
        cell.font = Font(bold=True)
    if centrado:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if borde:
        cell.border = _BRD
    if fondo:
        cell.fill = PatternFill("solid", fgColor=fondo)
    return cell


# ── Hoja de resumen ───────────────────────────────────────────────────────────

def _hoja_resumen(wb: Workbook, df_res: pd.DataFrame):
    ws = wb.active
    ws.title = "Resumen"
    hdrs = [
        "Sector", "Hileras", "Largo total (m)", "Largo prom (m)",
        "Largo min (m)", "Largo max (m)", "N° plantas",
        "Centrales", "Cent. Adic.", "Carpas", "Uso C (m²)",
        "Trans. totales", "Uso T (m)", "Uso P (m)",
    ]
    for ci, h in enumerate(hdrs, 1):
        _c(ws, 1, ci, h, negrita=True, centrado=True, borde=True, fondo=_AZUL)

    for ri, (_, row) in enumerate(df_res.iterrows(), 2):
        vals = [
            row["Sector"], row["Hileras"], row["Largo_total_m"], row["Largo_prom_m"],
            row["Largo_min_m"], row["Largo_max_m"], row["N_plantas_total"],
            row["Centrales_total"], row["Cent_Adic_total"], row["Carpas_total"],
            row["Uso_C_total"], row["Trans_total"], row["Uso_T_total"], row["Uso_P_total"],
        ]
        for ci, val in enumerate(vals, 1):
            _c(ws, ri, ci, val, borde=True)

    for ci in range(1, len(hdrs) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 16


# ── Hoja por sector ───────────────────────────────────────────────────────────

def _hoja_sector(wb: Workbook, sector: str, df_sec: pd.DataFrame,
                 tot: pd.Series, params: tuple,
                 d_hil_sec: float | None = None, d_pl_sec: float | None = None):
    ws = wb.create_sheet(title=sector[:31])

    (empresa, rut, encargado, dir_c, mail, celular,
     especie, variedad, sup_ha, alt_pl,
     d_hil, d_pl, m_hil, m_trans,
     ancho_c, l_carpa, l_min, alto_p, l_ent, alto_h, caida,
     ancho_vent, l_trans) = params

    # Rangos de la tabla — calculados antes para usarlos en fórmulas del resumen
    ini  = 16
    ft   = ini + len(df_sec)   # fila de totales
    last = ft - 1              # última fila de datos

    # ── Bloque cliente ────────────────────────────────────────────────────────
    for ref, txt in [("E2","Empresa:"),("E3","RUT:"),("E4","Dirección:"),("E5","Ciudad:"),
                     ("G2","Encargado:"),("G3","Mail:"),("G4","Celular:")]:
        _c(ws, ref, valor=txt, negrita=True)
    for ref, val in [("F2",empresa),("F3",rut),("F4",dir_c),
                     ("H2",encargado),("H3",mail),("H4",celular)]:
        ws[ref] = val

    # ── Resumen de alambres (N-O, filas 1-5) — fórmulas que apuntan a totales ──
    # Referencias a la fila de totales (ft):
    #   E{ft} = Largo total   G{ft} = Uso C total
    #   K{ft} = Uso P total   P{ft} = Uso T total
    _c(ws, "N1", valor="Tipo de alambre", negrita=True)
    _c(ws, "O1", valor="Largo (m)",       negrita=True)
    for fila, nombre in [(2,"CUMBRERAS"),(3,"PERIMETRO"),(4,"HOMBROS"),(5,"TRANS.")]:
        _c(ws, fila, 13, nombre, negrita=True)
    ws.cell(2, 15, f"=ROUND(E{ft}*(1+$G$8/100),2)")          # cumbreras
    ws.cell(3, 15, f"=ROUND(K{ft}+MAX(E{ini}:E{last})*2*(1+$G$8/100),2)")  # perímetro
    ws.cell(4, 15, f"=ROUND(E{ft}*2*(1+$G$8/100),2)")        # hombros
    ws.cell(5, 15, f"=ROUND(P{ft},2)")                        # transversales

    # ── Parámetros carpa (Q-S, filas 2-10) ───────────────────────────────────
    # R2=ancho_c  R3=l_carpa  R4=l_min  R5=alto_p  R6=l_ent
    # R7=alto_h   R8=caida    R9=ancho_vent  R10=l_trans
    _c(ws, "Q1", valor="Parámetro", negrita=True)
    _c(ws, "R1", valor="Valor",     negrita=True)
    _c(ws, "S1", valor="Unidad",    negrita=True)
    carpa_params = [
        ("Ancho carpa",       ancho_c,   "m"),
        ("Largo carpa",       l_carpa,   "m"),
        ("Largo mínimo",      l_min,     "m"),
        ("Alto pilares",      alto_p,    "m"),
        ("Largo enterrado",   l_ent,     "m"),
        ("Alto hombros",      alto_h,    "m"),
        ("Caída agua",        caida,     "m"),
        ("Ancho ventilación", ancho_vent,"m"),
        ("Largo transversal", l_trans,   "m"),   # R10 — referenciado en fórmulas
    ]
    for fi, (nombre, valor, unidad) in enumerate(carpa_params, 2):
        ws.cell(fi, 17, nombre).font = Font(bold=True)
        ws.cell(fi, 18, valor)
        ws.cell(fi, 19, unidad)

    # ── Datos del cultivo ─────────────────────────────────────────────────────
    # G8 y G9 se guardan como número (%, e.g. 5 para 5%) para que las fórmulas
    # puedan operar sobre ellos con /100.
    for ref, txt in [("A7","Especie:"),("A8","Variedad:"),
                     ("A9","Superficie (Ha):"),("A10","Altura de plantas (m):"),
                     ("D7","Marco plantación"),("D8","Entre hileras:"),
                     ("D9","Entre plantas:"),("F8","Merma Hileras (%)"),
                     ("F9","Merma Transversales (%)")]:
        _c(ws, ref, valor=txt, negrita=True)
    _d_hil = d_hil_sec if d_hil_sec is not None else d_hil
    _d_pl  = d_pl_sec  if d_pl_sec  is not None else d_pl
    for ref, val in [("B7",especie),("B8",variedad),("B9",sup_ha),("B10",alt_pl),
                     ("E8",_d_hil),("E9",_d_pl),("G8",m_hil),("G9",m_trans)]:
        ws[ref] = val

    # ── Encabezados de tabla ──────────────────────────────────────────────────
    ws["A12"] = f"Cuartel: {sector}"; ws["A12"].font = Font(bold=True)
    ws["A13"] = "Dist. centrales:";   ws["A13"].font = Font(bold=True)
    ws["B13"] = l_carpa   # ← referenciado como $B$13 en fórmulas de Centrales
    for ref, txt in [("I12","PERIMETRO"),("M12","TRANSVERSAL"),
                     ("I14","Total"),("M14","Total")]:
        _c(ws, ref, valor=txt, negrita=True)

    COL_HDRS = [
        "N° hilera","N° plantas","Centrales","Cent. Adic",
        "Largo (m)","Carpas","Uso C (m²)","",
        "Perimetro","Dist. hil","Uso P (m)","",
        "Transversal","hileras","Largo","Uso T (m)",
    ]
    for ci, h in enumerate(COL_HDRS, 1):
        if h:
            _c(ws, 15, ci, h, negrita=True, centrado=True, borde=True, fondo=_AZUL)

    # ── Filas de datos — fórmulas Excel ───────────────────────────────────────
    # Referencias absolutas de parámetros:
    #   $E$9  = dist_plantas       $B$13 = largo_carpa
    #   $G$8  = merma hileras (%)  $G$9  = merma transversales (%)
    #   $R$2  = ancho_carpa        $R$10 = largo_transversal
    for offset, (_, row) in enumerate(df_sec.iterrows()):
        fi = ini + offset
        E  = f"E{fi}"   # Largo_m
        C  = f"C{fi}"   # Centrales
        M  = f"M{fi}"   # Trans_cant

        vals = [
            int(row["N_hilera"]),                          # A  valor fijo
            f"=ROUND({E}/$E$9,2)",                         # B  N_plantas
            f"=INT({E}/$B$13)",                            # C  Centrales
            2,                                              # D  Cent. Adic (constante)
            row["Largo_m"],                                 # E  Largo medido (dato base)
            f"=IF({C}>0,{C}+1,0)",                        # F  Carpas
            f"=ROUND({E}*$R$2,2)",                        # G  Uso C m²
            "",                                             # H  separador
            2,                                              # I  Perim cant (constante)
            "=$E$8",                                        # J  Dist. hileras
            "=ROUND(2*$E$8*(1+$G$8/100),2)",              # K  Contribución lado corto
            "",                                             # L  separador
            f"=IF({C}>0,{C}+2,0)",                        # M  Trans cant
            1,                                              # N  constante
            "=$R$10",                                       # O  Largo transversal
            f"=ROUND({M}*$R$10*(1+$G$9/100),2)",         # P  Uso T m
        ]
        for ci, val in enumerate(vals, 1):
            if val != "":
                _c(ws, fi, ci, val, borde=True)

    # ── Fila de totales — fórmulas SUM ────────────────────────────────────────
    totales_fila = [
        "TOTAL",
        f"=ROUND(SUM(B{ini}:B{last}),0)",      # N_plantas total
        f"=SUM(C{ini}:C{last})",                # Centrales total
        f"=SUM(D{ini}:D{last})",                # Cent. Adic total
        f"=ROUND(SUM(E{ini}:E{last}),2)",      # Largo total
        f"=SUM(F{ini}:F{last})",                # Carpas total
        f"=ROUND(SUM(G{ini}:G{last}),2)",      # Uso C total
        "",
        f"=COUNTA(A{ini}:A{last})*2",           # Perim cant total
        "",
        f"=ROUND(SUM(K{ini}:K{last}),2)",      # Uso P total
        "",
        f"=SUM(M{ini}:M{last})",                # Trans cant total
        f"=COUNTA(A{ini}:A{last})",             # N° hileras
        "=$R$10",                                # Largo transversal
        f"=ROUND(SUM(P{ini}:P{last}),2)",      # Uso T total
    ]
    for ci, val in enumerate(totales_fila, 1):
        if val != "":
            _c(ws, ft, ci, val, negrita=True, borde=True, fondo=_AMAR)

    # ── Anchos de columna ─────────────────────────────────────────────────────
    for i, w in enumerate([10,12,11,12,10,8,12,3,13,10,12,3,14,9,10,12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Punto de entrada público ──────────────────────────────────────────────────

def crear_excel(df_det: pd.DataFrame, df_res: pd.DataFrame, params: tuple,
                params_por_sector: dict | None = None) -> bytes:
    wb = Workbook()
    _hoja_resumen(wb, df_res)

    for sector in df_res["Sector"].unique():
        df_s = df_det[df_det["Sector"] == sector].copy()
        df_s["N_hilera"] = range(1, len(df_s) + 1)
        tot = df_res[df_res["Sector"] == sector].iloc[0]
        d_hil_s, d_pl_s = (params_por_sector or {}).get(sector, (None, None))
        _hoja_sector(wb, sector, df_s, tot, params, d_hil_s, d_pl_s)

    wb.calculation.fullCalcOnLoad = True
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
